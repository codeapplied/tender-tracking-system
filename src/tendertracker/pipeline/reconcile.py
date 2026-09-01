import logging
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ..integrations.pipedrive import PipedriveClient
from ..storage.db import get_engine, get_session_factory
from ..storage.models import Tender

if TYPE_CHECKING:
    from ..config import Settings

logger = logging.getLogger(__name__)


@dataclass
class Discrepancy:
    external_id: str
    field: str
    db_value: str
    other_value: str
    other_system: str  # "excel" | "pipedrive"


@dataclass
class ReconcileResult:
    checked: int = 0
    discrepancies: list[Discrepancy] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    # What was actually available to compare against — "no discrepancies"
    # only means something was verified if these are true. Both false is a
    # real, distinct outcome: nothing was checked, not "everything's fine."
    excel_available: bool = False
    pipedrive_configured: bool = False


def _read_excel_statuses(path: str) -> tuple[dict[str, str], bool]:
    """Returns ({external_id: status}, available) read directly from the
    exported .xlsx — the one field the spreadsheet's dropdown lets a human
    edit. Everything else in Excel is a display copy of the DB, not an
    independent source. `available` is False if the file doesn't exist yet
    (never exported) or has no sheet with the expected columns — distinct
    from "available but genuinely empty," which is a real, comparable state."""
    statuses: dict[str, str] = {}
    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        return statuses, False

    found_expected_columns = False
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if "External ID" not in headers or "Status" not in headers:
            continue
        found_expected_columns = True
        id_idx = headers.index("External ID")
        status_idx = headers.index("Status")
        for row in ws.iter_rows(min_row=2, values_only=True):
            external_id = row[id_idx]
            if external_id:
                statuses[external_id] = row[status_idx]
    return statuses, found_expected_columns


def reconcile(settings: "Settings") -> ReconcileResult:
    """Read-only. Compares DB vs. Excel (status — the one human-editable
    field the spreadsheet exposes) and DB vs. Pipedrive (title, value —
    the fields sync_pipedrive writes and a salesperson could independently
    edit in the Pipedrive UI). Reports differences, never writes a fix —
    matches the old system's own governance pattern: surface drift for a
    human to resolve, don't auto-resolve it.
    """
    result = ReconcileResult()
    excel_statuses, result.excel_available = _read_excel_statuses(settings.excel_export_path)

    pipedrive_client = None
    if settings.pipedrive_api_token and settings.pipedrive_domain:
        pipedrive_client = PipedriveClient(settings.pipedrive_api_token, settings.pipedrive_domain)
        result.pipedrive_configured = True

    engine = get_engine(settings.db_path)
    session_factory = get_session_factory(engine)
    with session_factory() as session:
        for tender in session.query(Tender).all():
            result.checked += 1

            excel_status = excel_statuses.get(tender.external_id)
            if excel_status is not None and excel_status != tender.status:
                result.discrepancies.append(
                    Discrepancy(tender.external_id, "status", tender.status, str(excel_status), "excel")
                )

            if pipedrive_client and tender.pipedrive_deal_id:
                try:
                    deal = pipedrive_client.get_deal(int(tender.pipedrive_deal_id))
                except Exception as exc:
                    result.errors.append(f"{tender.external_id}: Pipedrive fetch failed: {exc}")
                    continue

                deal_title = deal.get("title")
                if deal_title is not None and deal_title != tender.title:
                    result.discrepancies.append(
                        Discrepancy(tender.external_id, "title", tender.title, str(deal_title), "pipedrive")
                    )

                db_value = tender.estimated_value
                deal_value = deal.get("value")
                # deal_value == 0 is treated as "not set," same as None — a
                # deliberate but unverified tradeoff: Pipedrive's API isn't
                # confirmed live in this project (no live account available,
                # see integrations/pipedrive.py), so it's not known whether
                # value=0 means "genuinely zero" or "field left blank." A
                # false negative (missing a real $0 deal) was judged safer
                # than a false positive on every newly-created, not-yet-priced
                # deal. Revisit against a real Pipedrive account if this
                # matters for your use case.
                if db_value is not None and deal_value not in (None, 0) and float(db_value) != float(deal_value):
                    result.discrepancies.append(
                        Discrepancy(tender.external_id, "estimated_value", str(db_value), str(deal_value), "pipedrive")
                    )

    return result
