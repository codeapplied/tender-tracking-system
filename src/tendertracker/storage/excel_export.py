from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from .models import Tender

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

COLUMNS = [
    ("source", "Source"),
    ("external_id", "External ID"),
    ("title", "Title"),
    ("category", "Category"),
    ("status", "Status"),
    ("relevance_score", "Relevance"),
    ("published_date", "Published"),
    ("closing_date", "Closing"),
    ("estimated_value", "Value"),
    ("currency", "Currency"),
    ("url", "URL"),
]

# Human-owned in the DB (see run_daily.HUMAN_OWNED_FIELDS) — the dropdown
# here is for humans editing the spreadsheet, not fed back from anywhere yet.
STATUS_OPTIONS = ["open", "reviewing", "bidding", "submitted", "won", "lost", "closed"]

CLOSED_STATUSES = {"closed", "lost"}


def _write_sheet(ws: Worksheet, tenders: list[Tender]) -> None:
    ws.append([label for _, label in COLUMNS])
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"

    for tender in tenders:
        ws.append([getattr(tender, field) for field, _ in COLUMNS])

    status_col_idx = next(i for i, (field, _) in enumerate(COLUMNS, start=1) if field == "status")
    status_col_letter = get_column_letter(status_col_idx)
    last_row = max(len(tenders) + 1, 2)
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{status_col_letter}2:{status_col_letter}{last_row}")

    ws.auto_filter.ref = ws.dimensions

    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(label) + 4)


def export_to_excel(session: Session, path: str) -> tuple[int, int]:
    """Regenerate the tracked .xlsx from current DB state.

    Idempotent — fully overwrites the file from the DB each call, no merge
    with prior file content (the DB is the source of truth; Excel is a
    read/edit surface synced from it, not the other way around — see
    HUMAN_OWNED_FIELDS in run_daily.py for what a human can still change
    that survives re-sync).

    Returns (active_count, archived_count).
    """
    tenders = session.query(Tender).order_by(Tender.closing_date.asc().nullslast()).all()
    active = [t for t in tenders if t.status not in CLOSED_STATUSES]
    archived = [t for t in tenders if t.status in CLOSED_STATUSES]

    wb = Workbook()
    wb.active.title = "Active"
    _write_sheet(wb.active, active)
    _write_sheet(wb.create_sheet("Archived"), archived)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    return len(active), len(archived)
