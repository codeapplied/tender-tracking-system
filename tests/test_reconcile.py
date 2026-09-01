from unittest.mock import MagicMock, patch

import pytest

from tendertracker.config import PortalConfig
from tendertracker.pipeline.reconcile import reconcile
from tendertracker.pipeline.run_daily import run_source
from tendertracker.storage.excel_export import export_to_excel
from tendertracker.storage.models import Tender

SANDBOX_PORTAL = PortalConfig(name="sandbox", scraper_class="tendertracker.scrapers.sandbox_feed.SandboxFeedScraper")


@pytest.fixture
def populated(tmp_settings, db_session):
    run_source(db_session, SANDBOX_PORTAL)
    db_session.commit()
    export_to_excel(db_session, tmp_settings.excel_export_path)
    return tmp_settings, db_session


def test_no_drift_on_fresh_export(populated):
    settings, _ = populated
    result = reconcile(settings)
    assert result.checked == 5
    assert result.discrepancies == []
    assert result.excel_available is True
    assert result.pipedrive_configured is False


def test_excel_unavailable_when_never_exported(tmp_settings, db_session):
    """Distinct from 'no discrepancies' — nothing to compare against yet,
    since export was never run. Regression test for a real gap: this used
    to be indistinguishable from a genuinely-verified clean result."""
    run_source(db_session, SANDBOX_PORTAL)
    db_session.commit()

    result = reconcile(tmp_settings)
    assert result.excel_available is False
    assert result.pipedrive_configured is False
    assert result.discrepancies == []  # would misleadingly look "clean" without the availability flags


def test_catches_status_edited_directly_in_excel(populated):
    from openpyxl import load_workbook

    settings, db_session = populated

    wb = load_workbook(settings.excel_export_path)
    ws = wb["Active"]
    headers = [c.value for c in ws[1]]
    status_col = headers.index("Status") + 1
    ws.cell(row=2, column=status_col).value = "bidding"
    wb.save(settings.excel_export_path)

    result = reconcile(settings)
    status_discrepancies = [d for d in result.discrepancies if d.field == "status"]
    assert len(status_discrepancies) == 1
    assert status_discrepancies[0].db_value == "open"
    assert status_discrepancies[0].other_value == "bidding"
    assert status_discrepancies[0].other_system == "excel"


def test_reconcile_never_writes_to_db(populated):
    from openpyxl import load_workbook

    settings, db_session = populated
    wb = load_workbook(settings.excel_export_path)
    ws = wb["Active"]
    headers = [c.value for c in ws[1]]
    status_col = headers.index("Status") + 1
    ws.cell(row=2, column=status_col).value = "bidding"
    wb.save(settings.excel_export_path)

    reconcile(settings)

    tender = db_session.query(Tender).first()
    assert tender.status == "open"  # unchanged by reconcile itself


def test_pipedrive_value_drift_detected(populated):
    settings, db_session = populated
    settings.pipedrive_api_token = "FAKE"
    settings.pipedrive_domain = "fakecompany"

    tender = db_session.query(Tender).first()
    tender.pipedrive_deal_id = "123"
    db_session.commit()

    with patch("tendertracker.pipeline.reconcile.PipedriveClient") as MockClient:
        MockClient.return_value.get_deal.return_value = {"title": tender.title, "value": 999999}
        result = reconcile(settings)

    value_discrepancies = [d for d in result.discrepancies if d.field == "estimated_value"]
    assert len(value_discrepancies) == 1
    assert value_discrepancies[0].other_system == "pipedrive"
    assert result.pipedrive_configured is True


def test_pipedrive_fetch_failure_recorded_as_error_not_crash(populated):
    settings, db_session = populated
    settings.pipedrive_api_token = "FAKE"
    settings.pipedrive_domain = "fakecompany"

    tender = db_session.query(Tender).first()
    tender.pipedrive_deal_id = "123"
    db_session.commit()

    with patch("tendertracker.pipeline.reconcile.PipedriveClient") as MockClient:
        MockClient.return_value.get_deal.side_effect = Exception("network error")
        result = reconcile(settings)

    assert len(result.errors) == 1
    assert "network error" in result.errors[0]
